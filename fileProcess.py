from math import ceil

import openpyxl
import numpy as np
# 将表格中的数据按故障类别分为7类
num_topics = 7

def xls2txt(sub_dic):
    wb = openpyxl.load_workbook("/Users/qiuhongzhai/Desktop/故障检测报告/故障数据_改2.xlsx")  # 工作簿(workbook)对象
    ws = wb.active  # 表单对象
    column = ws['A']  # A列是故障类别
    for c in column:
        if c.row == 1:  # 跳过第一行（名字）
            continue
        try:
            if c._value >= 1 and c._value <= num_topics:
                path = "Data/" + sub_dic + "/" + "class_" + str(c._value) + ".txt"
                with open(path, "a") as f:  # 为了减少OS成本，改成分别拼接各个类的字符串，再分别一次写入会比较好
                    pos = "B" + str(c.row)
                    f.write(ws[pos]._value.replace(" ", "").replace("\n", "") + '\n')
            else:  # 类别出错
                print("row : " + c.row + "class error!")
                continue
        except:
            print(c._value)


def aggregate(sub_dic):
    content = ""
    for i in range(7):
        path = "Data/class_" + sub_dic + "/" + str(i+1) + ".txt"
        with open(path, "r") as f:
            for line in f:
                # 判断这一行是否为空字符串
                line = line.replace(" ", "").replace("\n", "")
                if line.__len__() == 0 :
                    continue
                content += line + '\n'
    with open("Data/" + sub_dic + "/" + "all_class.txt","w") as outFile:
        outFile.write(content)

# shuffle dataset
def shuffle(sub_dic):
    for i in range(7):
        path = "Data/" + sub_dic + "/" + "class_" + str(i + 1) + ".txt"
        count = 0
        content = []
        new_content = []
        with open(path, "r") as f:
            for line in f:
                # 判断这一行是否为空字符串
                line = line.replace(" ", "").replace("\n", "")
                if line.__len__() == 0 :
                    continue
                count = count + 1
                content.append(line)
            arr = np.arange(count)
            np.random.shuffle(arr)  # 随机打乱arr数组
            for index in arr:
                new_content.append(content[index])
            out_str = ""
            for con in new_content:
                out_str = out_str + str(i+1) + "\t" + con + "\n"
            shuffled_path = "Data/" + sub_dic + "/" + "class_" + str(i + 1) + "_shuffled.txt"
            with open(shuffled_path,"w") as shu_f:
                shu_f.write(out_str)


# 将数据集分成7:3
def cutDataSet(sub_dic):
    train_path = "Data/" + sub_dic + "/" + "train.txt"
    test_path = "Data/" + sub_dic + "/" + "test.txt"
    for i in range(7):
        count = 0
        docs = []
        path = "Data/" + sub_dic + "/" + "class_" + str(i + 1) + "_shuffled.txt"
        with open(path, "r") as f:
            for line in f:
                line = line.replace(" ", "").replace("\n", "")
                if line.__len__() == 0:
                    continue
                docs.append(line)
                count = count + 1
        train_num = ceil(count * 0.7)
        train_str = ""
        test_str = ""
        for j in range(train_num):
            train_str = train_str + docs[j]+"\n"

        for m in range(train_num, count):
            test_str = test_str + docs[m]+"\n"

        with open(train_path, "a") as train_file:
            train_file.write(train_str)
        with open(test_path, "a") as test_file:
            test_file.write(test_str)

def addLabel():   # 临时函数 用于给test文档加上对应的label
    test_num = []
    for i in range(7):
        count = 0
        path = "Data/class_" + str(i + 1) + ".txt"
        with open(path, "r") as f:
            for line in f:
                line = line.replace(" ", "").replace("\n", "")
                if line.__len__() == 0:
                    continue
                count = count + 1
        test_num.append(count - ceil(count * 0.7))

    test_path = "Data/test.txt"
    test_str = []
    with open(test_path, "r") as test_file_r:
        for line in test_file_r:
            line = line.replace(" ", "").replace("\n", "")
            if line.__len__() == 0:
                continue
            test_str.append(line)
    label = 1
    test_index = 0
    new_test_str = ""
    for num in test_num:  # 7类各自的test条数
        for i in range(num):  # 打印num次某label，label初始等于1，后面累加
            # test_file_w.write(str(label) + "\t" + test_str[test_index] + "\n")
            new_test_str = new_test_str + (str(label) + "\t" + test_str[test_index] + "\n")
            test_index = test_index + 1
        label = label + 1
    with open(test_path, "w") as test_file_w:
        test_file_w.write(new_test_str)



if __name__ == '__main__':
    sub_dic = "second"
    xls2txt(sub_dic)
    shuffle(sub_dic)
    # aggregate()     # 将7类的txt合成一个
    cutDataSet(sub_dic)
    # addLabel()
